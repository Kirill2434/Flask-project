from openpyxl import load_workbook
from flask import Blueprint, Flask, request
from main import excel_validation_bp, BasicTypeGraphic, CheckedTypeGraphic, db


@excel_validation_bp.route("/get_file", methods=["POST"])
def making_dict():
    file = request.files['file']
    wb = load_workbook(file)
    sheet = wb.get_sheet_by_name('График ЗП')
    data = sheet.iter_rows(min_row=4, max_row=4, min_col=5, max_col=35, values_only=True)
    hours = sheet.iter_rows(min_row=5, max_row=5, min_col=5,  max_col=35, values_only=True)
    dictionary = dict(zip(*data, *hours))

    def upd_to_float_dict_1(dict_d):
        upd_dict_1 = {}
        for key, value in dict_d.items():
            new_value = None

            if value == "'" or value == "":
                new_value = 0

            if new_value != 0:
                new_value = ''
                for i in str(value):
                    if i == ',':
                        i = '.'
                    new_value = new_value + i
            upd_dict_1[str(key)] = float(new_value)

        return upd_dict_1
    # не видит первый словарь
    correct_dictionary_base_file = upd_to_float_dict_1(dictionary)
    main_base = BasicTypeGraphic(main_date=data, hours=hours)
    db.session.add(main_base)
    db.session.commit()
    return {"ok": True}


@excel_validation_bp.route("/get_file_check", methods=["POST"])
def get_hours_list():
    file = request.files['file2']
    wb = load_workbook(file)
    sheet = wb.active
    hours_list = []
    for row in range(22, 632, 4):
        hours_dict = {}
        date_1 = sheet.iter_rows(min_row=14, max_row=14, min_col=4, max_col=18, values_only=True)
        date_2 = sheet.iter_rows(min_row=18, max_row=18, min_col=4, max_col=19, values_only=True)
        hours_1 = sheet.iter_rows(min_row=row, max_row=row, min_col=4, max_col=18, values_only=True)
        hours_2 = sheet.iter_rows(min_row=row + 2, max_row=row + 2, min_col=4, max_col=19, values_only=True)
        dictionary = dict(zip(*date_1, *hours_1))
        dictionary_2 = dict(zip(*date_2, *hours_2))
        hours_dict['dictionary'] = dictionary
        hours_dict['dictionary_2'] = dictionary_2
        hours_list.append(hours_dict)
    return hours_list

    hours_list = get_hours_list()

    def upd_to_float_dict_2(dictionary1):
        upd_dict = {}
        for key, value in dictionary1.items():
            new_value = None

            if value == 'X':
                continue

            if value == "'" or value == "":
                new_value = 0

            if new_value != 0:
                new_value = ''
                for i in str(value):
                    if i == ',':
                        i = '.'
                    new_value = new_value + i
            upd_dict[str(key)] = float(new_value)
        return upd_dict

    result_dict = {}
    for hours_dict3 in hours_list:
        upd_dict_1 = upd_to_float_dict_2(hours_dict3['dictionary'])
        upd_dict_2 = upd_to_float_dict_2(hours_dict3['dictionary_2'])
        result_dict.update(upd_dict_1)
        result_dict.update(upd_dict_2)

    # добавление в базу
        check_base = CheckedTypeGraphic(main_date=data, hours=hours)
        db.session.add(check_base)
        db.session.commit()

    # Сравниваем  словари
    def comp_dicts(base_dict, check_dict):
        for key, val in base_dict.items():
            if val == check_dict[key]:
                print('Ok', val, check_dict[key])
            else:
                print('Не совпало элементов:', val, check_dict[key])
        return base_dict == check_dict

    print(comp_dicts(correct_dictionary_base_file, result_dict))

    return {"ok": True}





