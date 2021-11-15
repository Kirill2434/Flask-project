from openpyxl import load_workbook
from flask import Blueprint, request
from app_R_K.excel_validation.utilits import upd_to_float_dict_1, upd_to_float_dict_2, comparison_of_dictionaries
from models import BasicTypeGraphic, db

validation_bp = Blueprint('validation_bp', __name__)


@validation_bp.route("/get_file", methods=["POST"])
def making_dict():

    file = request.files['file']
    wb = load_workbook(file)
    sheet = wb.get_sheet_by_name('График ЗП')
    data = sheet.iter_rows(min_row=4, max_row=4, min_col=5, max_col=35, values_only=True)
    hours = sheet.iter_rows(min_row=5, max_row=5, min_col=5,  max_col=35, values_only=True)
    dictionary = dict(zip(*data, *hours))
    correct_dictionary_base_file = upd_to_float_dict_1(dictionary)
    print(correct_dictionary_base_file)
    # добавление в базу
    #TODO: Сделать проверку на существование в базе такой записи

    for key, value in correct_dictionary_base_file.items():
        main_base = BasicTypeGraphic(main_date=key, hours=value)
        db.session.add(main_base)

    db.session.commit()

    return {"ok": True}


surname_list = []


@validation_bp.route("/get_file_check", methods=["POST"])
def get_hours_list():
    file = request.files['file']
    wb = load_workbook(file)
    sheet = wb.get_sheet_by_name('Лист1')
    basic_record_from_base = BasicTypeGraphic.query.with_entities(BasicTypeGraphic.main_date, BasicTypeGraphic.hours).all()

    basic_record_from_base_dict = {}
    for key, value in basic_record_from_base:
        basic_record_from_base_dict[key] = float(value)

    hours_list = []
    for row in range(22, 632, 4):
        hours_dict = {}

        date_1 = sheet.iter_rows(min_row=14, max_row=14, min_col=4, max_col=18, values_only=True)
        date_2 = sheet.iter_rows(min_row=18, max_row=18, min_col=4, max_col=19, values_only=True)
        hours_1 = sheet.iter_rows(min_row=row, max_row=row, min_col=4, max_col=18, values_only=True)
        hours_2 = sheet.iter_rows(min_row=row + 2, max_row=row + 2, min_col=4, max_col=19, values_only=True)

        dictionary = dict(zip(*date_1, *hours_1))
        dictionary_2 = dict(zip(*date_2, *hours_2))

        hours_dict['dictionary'] = dictionary
        hours_dict['dictionary_2'] = dictionary_2
        hours_list.append(hours_dict)
        result_dict = {}
        for hours_dict in hours_list:
            upd_dict_1 = upd_to_float_dict_2(hours_dict['dictionary'])
            upd_dict_2 = upd_to_float_dict_2(hours_dict['dictionary_2'])
            result_dict.update(upd_dict_1)
            result_dict.update(upd_dict_2)

        for row_s in range(21, 629, 4):
            surname = sheet.cell(row=row_s + 4, column=2).value
            surname_list.append(surname)

        print(comparison_of_dictionaries(basic_record_from_base_dict, result_dict))

    return {'result': hours_list}
